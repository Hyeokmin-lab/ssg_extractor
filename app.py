"""
SSG.COM 상품 정보 추출기 - Streamlit UI
"""

import streamlit as st
import pandas as pd
from io import BytesIO
import os
import zipfile
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

from ssg_product_extractor import create_driver, extract_product

# ────────────────────────────────────────────────────────────────
# 페이지 설정
# ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SSG.COM 상품 정보 추출기",
    page_icon="🛒",
    layout="centered",
)

# ────────────────────────────────────────────────────────────────
# 스타일 (롯데ON 추출기와 동일한 레이아웃 기준 720px)
# ────────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
    /* 중앙 720px 레이아웃 */
    .main .block-container {
        max-width: 720px;
        padding-top: 2rem;
        padding-bottom: 2rem;
    }

    /* 타이틀 */
    h1 { font-size: 1.8rem !important; font-weight: 700 !important; }

    /* URL 입력 박스 */
    textarea {
        font-size: 0.875rem !important;
        background-color: #f8f9fa !important;
        border-radius: 6px !important;
    }

    /* 버튼 */
    .stButton > button {
        height: 48px;
        font-size: 0.95rem;
        font-weight: 600;
        border-radius: 6px;
    }

    /* 추출 시작 버튼 (primary) */
    .stButton > button[kind="primary"] {
        background-color: #e74c3c !important;
        border-color: #e74c3c !important;
    }
    .stButton > button[kind="primary"]:hover {
        background-color: #c0392b !important;
        border-color: #c0392b !important;
    }

    /* 진행 상태 */
    .status-box {
        background: #f0f4ff;
        border-left: 4px solid #3b82f6;
        padding: 10px 14px;
        border-radius: 4px;
        font-size: 0.875rem;
        color: #1e40af;
        margin: 8px 0;
    }

    /* 결과 테이블 */
    .stDataFrame { border-radius: 8px; overflow: hidden; }

    /* 섹션 구분선 */
    hr { margin: 1.2rem 0 !important; }

    /* 다운로드 버튼 */
    .stDownloadButton > button {
        background-color: #22c55e !important;
        color: white !important;
        border-color: #22c55e !important;
        height: 44px;
        font-weight: 600;
        border-radius: 6px;
        width: 100%;
    }
</style>
""",
    unsafe_allow_html=True,
)

# ────────────────────────────────────────────────────────────────
# 세션 상태 초기화
# ────────────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []
if "log_msgs" not in st.session_state:
    st.session_state.log_msgs = []
if "image_dir" not in st.session_state:
    st.session_state.image_dir = None

# ────────────────────────────────────────────────────────────────
# 헤더
# ────────────────────────────────────────────────────────────────
st.title("🛒 SSG.COM 상품 정보 추출기")
st.markdown("---")

# ────────────────────────────────────────────────────────────────
# URL 입력
# ────────────────────────────────────────────────────────────────
st.markdown("🔗 **SSG.COM 상품 URL 입력**")
urls_input = st.text_area(
    label="urls",
    label_visibility="collapsed",
    placeholder=(
        "URL을 한 줄에 하나씩 입력하세요.\n"
        "(예: https://www.ssg.com/item/itemView.ssg?itemId=1000795246042)"
    ),
    height=160,
    key="url_textarea",
)

# ────────────────────────────────────────────────────────────────
# 이미지 저장 옵션
# ────────────────────────────────────────────────────────────────
with st.expander("📷 상품상세 이미지 저장 옵션", expanded=False):
    save_images = st.checkbox("상품상세 이미지 다운로드", value=False, key="save_images")
    image_save_path = st.text_input(
        "저장 폴더 경로",
        value=os.path.join(os.path.expanduser("~"), "Desktop", "ssg_images"),
        help="이미지를 저장할 로컬 폴더 경로 (모델번호별 하위 폴더가 자동 생성됩니다)",
        disabled=not save_images,
        key="image_save_path",
    )

st.markdown("---")

# ────────────────────────────────────────────────────────────────
# 버튼 행
# ────────────────────────────────────────────────────────────────
col_start, col_reset = st.columns([2, 1])
with col_start:
    start_btn = st.button("🔍 추출 시작", use_container_width=True, type="primary")
with col_reset:
    reset_btn = st.button("🔄 초기화", use_container_width=True)

# 초기화
if reset_btn:
    st.session_state.results = []
    st.session_state.log_msgs = []
    st.rerun()

# ────────────────────────────────────────────────────────────────
# 추출 실행
# ────────────────────────────────────────────────────────────────
if start_btn:
    raw_urls = [u.strip() for u in urls_input.strip().splitlines() if u.strip()]
    if not raw_urls:
        st.warning("URL을 한 줄 이상 입력해주세요.")
    else:
        st.session_state.results = []
        st.session_state.log_msgs = []

        prog_bar    = st.progress(0)
        status_box  = st.empty()
        log_area    = st.empty()

        results  = []
        log_msgs = []

        def update_log(msg):
            log_msgs.append(msg)
            log_area.markdown(
                "\n".join(f"- {m}" for m in log_msgs[-5:]),
                unsafe_allow_html=True,
            )

        driver = create_driver(headless=True)
        try:
            for idx, url in enumerate(raw_urls):
                status_box.markdown(
                    f'<div class="status-box">⏳ 처리 중... '
                    f'({idx + 1} / {len(raw_urls)})  '
                    f'<code>{url[:70]}</code></div>',
                    unsafe_allow_html=True,
                )
                update_log(f"[{idx+1}] {url[:60]}... 추출 시작")

                img_dir = image_save_path if save_images else None
                res = extract_product(driver, url, status_callback=update_log, image_dir=img_dir)
                results.append(res)

                if res["오류"]:
                    update_log(f"  ❌ 오류: {res['오류'][:60]}")
                else:
                    update_log(
                        f"  ✅ 완료 | {res['브랜드']} | {res['상품명'][:30]}"
                    )

                prog_bar.progress((idx + 1) / len(raw_urls))
                time.sleep(0.5)

        finally:
            driver.quit()

        st.session_state.results  = results
        st.session_state.log_msgs = log_msgs
        st.session_state.image_dir = image_save_path if save_images else None

        status_box.markdown(
            f'<div class="status-box" style="border-color:#22c55e;background:#f0fdf4;color:#15803d;">'
            f"✅ 완료! 총 {len(results)}개 상품 추출 완료</div>",
            unsafe_allow_html=True,
        )

# ────────────────────────────────────────────────────────────────
# 결과 표시
# ────────────────────────────────────────────────────────────────
if st.session_state.results:
    st.markdown("---")
    st.subheader(f"📋 추출 결과 ({len(st.session_state.results)}개)")

    df = pd.DataFrame(st.session_state.results)
    display_cols = ["브랜드", "상품명", "판매가", "색상", "사이즈", "모델번호", "오류"]
    available_cols = [c for c in display_cols if c in df.columns]
    st.dataframe(df[available_cols], use_container_width=True, height=300)

    # ── Excel 생성 ──────────────────────────────────────────────
    def build_excel(results):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SSG 상품정보"

        # 헤더 정의
        headers = [
            "번호", "브랜드", "상품명", "판매가", "색상",
            "사이즈", "모델번호", "URL", "상품상세이미지URL", "오류",
        ]
        col_widths = [6, 16, 50, 12, 30, 30, 16, 50, 60, 30]

        # 헤더 스타일
        header_fill  = PatternFill("solid", fgColor="C0392B")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws.row_dimensions[1].height = 28
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 데이터 행
        alt_fill = PatternFill("solid", fgColor="FEF9F9")
        for row_idx, res in enumerate(results, start=2):
            ws.row_dimensions[row_idx].height = 20
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            row_data = [
                row_idx - 1,
                res.get("브랜드", ""),
                res.get("상품명", ""),
                res.get("판매가", ""),
                res.get("색상", ""),
                res.get("사이즈", ""),
                res.get("모델번호", ""),
                res.get("URL", ""),
                res.get("상품상세이미지", ""),
                res.get("오류", ""),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = thin_border
                cell.alignment = center_align if col_idx == 1 else left_align
                if fill.fill_type:
                    cell.fill = fill

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    excel_buf = build_excel(st.session_state.results)

    st.markdown("&nbsp;")
    dl_col1, dl_col2 = st.columns(2)

    with dl_col1:
        st.download_button(
            label="📥 Excel 다운로드",
            data=excel_buf,
            file_name="ssg_products.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with dl_col2:
        # 이미지 폴더가 있으면 ZIP으로 묶어 다운로드 제공
        img_dir = st.session_state.get("image_dir")
        if img_dir and os.path.isdir(img_dir):
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(img_dir):
                    for file in files:
                        abs_path = os.path.join(root, file)
                        arc_name = os.path.relpath(abs_path, os.path.dirname(img_dir))
                        zf.write(abs_path, arc_name)
            zip_buf.seek(0)
            st.download_button(
                label="🖼️ 상세이미지 ZIP 다운로드",
                data=zip_buf,
                file_name="ssg_detail_images.zip",
                mime="application/zip",
                use_container_width=True,
            )
        else:
            st.button(
                "🖼️ 상세이미지 ZIP 다운로드",
                disabled=True,
                use_container_width=True,
                help="이미지 저장 옵션을 활성화한 후 추출하면 다운로드가 가능합니다.",
            )
